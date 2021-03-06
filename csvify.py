#/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import errno
import glob
import csv
import openpyxl
	
import argparse
from itertools import tee

SUBMODULES = ["python-docx"]
TBL_HEADER_MAX_SIZE = 2

XLSX_MAX_COLUMNS = 200
XLSX_HEADER_MAX_SIZE = 200

def _module_path():
	if "__file__" in globals():
		return os.path.dirname(os.path.realpath(__file__))
	
	return ""
		
def setup_env():
	global SUBMODULES
	
	this_path = _module_path()
	SUBMODULES = [os.path.join(this_path, sub) for sub in SUBMODULES]
	sys.path = SUBMODULES + sys.path

setup_env()

import docx

def mkdir_p(path):
	try:
		os.makedirs(path)
	except OSError as exc:  # Python >2.5
		if exc.errno == errno.EEXIST and os.path.isdir(path):
			pass
		else:
			raise
			
def parse_docx_tbl(tbl, keep_header = False, header_size = -1, keep_newlines = False):
	"""
	Parse a python-docx table object *tbl* and return
	the data in csv-compatible format.
	"""
	res = []
	detect_header_size = (header_size == -1)
	header_size = 0

	if detect_header_size and not keep_header:
		header_size = 1

		if len(tbl.rows) < 2:
			return [[]]

		if (tbl.rows[0].cells[0]._tc == tbl.rows[1].cells[0]._tc):
			header_size = TBL_HEADER_MAX_SIZE

	for nrow, row in enumerate(tbl.rows[header_size:]):
		last_tc = None
		buf = []
		for cell in row.cells:
			# ignore merged and empty cells
			if (cell._tc != last_tc):
				text = cell.text
				if (not keep_newlines):
					text = cell.text.replace("\r\n", " ") \
							.replace("\n", " ")
				buf.append(text)
			last_tc = cell._tc
		res.append(buf)
	return res

def write_csv(lines, output_file):
	"""
	Write *lines* to *output_file*.
	"""
	with open(output_file, "w") as f:
		writer = csv.writer(f)
		writer.writerows(lines)
		
def docx_tbl2csv(tbl, output_file, keep_header = False, header_size = 1, keep_newlines = False):
	"""
	Parse a python-docx table object *tbl* and write
	the result as *output_file* in CSV format.
	"""
	write_csv(parse_docx_tbl(tbl, keep_header, header_size, keep_newlines), output_file)
	
def parse_named_range(wb, named_range):
	"""
	Get a single rectangular region from
	the specified openpyxl workbook object *wb*.
	*named_range* is openpyxl.workbook.defined_name.DefinedName object.
	"""
	
	range_name = named_range.name
	
	# skip cell references
	if '!' in range_name:
		return [[]]

	destinations = list(named_range.destinations) 
	ws, reg = destinations[0]
	ws = wb[ws]
	region = ws[reg]

	for row in region:
		yield row

def worksheet2iter(ws):
	for row in ws:
		yield row

def read_row(ws, row_num):
	for i in range(1, XLSX_MAX_COLUMNS + 1):
		yield ws.cell(row = row_num, column = i).value

def xlsx_region2csv(ws, region_iter, output_file, keep_header = False, header_size = 0, keep_newlines = False):
	"""
	Iterate through *region_iter* and write
	the result as *output_file* in CSV format.
	"""
	
	reg = []
	
	def __process_row(row):
		if not row:
			return

		if (keep_newlines):
			row_ = [cell.value for cell in row]
		else:
			row_ = [cell.value.replace("\r\n", " ").replace("\n", " ") \
					if isinstance(cell.value, str) else cell.value for cell in row]

		# return early if encountered an empty row
		if not any(row_):
			return
		reg.append(row_)

	first_row = []

	# skip empty rows at the beginning of the sheet
	for _ in range(XLSX_HEADER_MAX_SIZE):
		try:
			first_row = next(region_iter)
			if any(cell.value for cell in first_row):
				#print("detected header size of %d rows" %i)
				break
		except StopIteration:
			break

	__process_row(first_row)

	first_row = []

	# optionally skip the header
	if not keep_header:
		for _ in range(header_size):
			try:
				first_row = next(region_iter)
			except StopIteration:
				break

		__process_row(first_row)

	for row in region_iter:
		__process_row(row)

	write_csv(reg, output_file)

def main(input_file, output_dir = "", use_captions = True,
			use_named_ranges = False, keep_header = False, header_size = 0, keep_newlines = False):
	if not os.path.exists(input_file):
		parser.error("Specified input file(folder) %s was not found" % input_file)
		return

	input_is_dir = os.path.isdir(input_file)

	if not output_dir:
		if input_is_dir:
			output_dir = os.path.join(input_file, "out")
		else:
			output_dir = os.path.join(os.path.dirname(input_file), "out")

		print(output_dir)

	if input_is_dir:
		old_pwd = os.getcwd()
		os.chdir(input_file)	
		in_docx_files = [i for i in glob.iglob("**/*.docx", recursive=True) if os.path.isfile(i)]
		in_xlsx_files = [i for i in glob.iglob("**/*.xlsx", recursive=True) if os.path.isfile(i)]
	else:
		in_docx_files = []
		in_xlsx_files = []
		if input_file.endswith(".docx"):
			in_docx_files = [input_file]
		elif input_file.endswith(".xlsx"):
			in_xlsx_files = [input_file]

	# main processing routine
	for docx_filename in in_docx_files:
		print("processing {file}".format(file = docx_filename))
		out_dir = os.path.join(output_dir, os.path.basename(docx_filename))
		mkdir_p(out_dir)
		tables = docx.Document(docx_filename).tables
		for n, tbl in enumerate(tables):
			if use_captions and hasattr(tbl, "caption") and tbl.caption:
				out_file = os.path.join(out_dir, 
					"{tbl_name}.csv".format(tbl_name = tbl.caption))
				print("found table {tbl_name}, saving as {out_file}"
					.format(tbl_name = tbl.caption, out_file = out_file))
			else:
				out_file = os.path.join(out_dir, 
					"{tbl_num}.csv".format(tbl_num = n))
				print("found table #{tbl_num}, saving as {out_file}"
					.format(tbl_num = n, out_file = out_file))

			docx_tbl2csv(tbl, out_file, keep_header, header_size, keep_newlines)

	for xlsx_filename in in_xlsx_files:
		print("processing {file}".format(file = xlsx_filename))
		out_dir = os.path.join(output_dir, os.path.basename(xlsx_filename))
		mkdir_p(out_dir)
		wb = openpyxl.load_workbook(xlsx_filename, data_only = True, read_only = True)
		if use_named_ranges:
			for named_range in wb.defined_names.definedName:
				out_file = os.path.join(out_dir, 
						"{reg_name}.csv".format(reg_name = named_range.name))
				xlsx_region2csv(named_range, parse_named_range(wb, named_range), out_file, keep_header, header_size, keep_newlines)

		for ws in wb.worksheets:
			out_file = os.path.join(out_dir, 
					"{ws_name}.csv".format(ws_name = ws.title))
			xlsx_region2csv(ws, worksheet2iter(ws), out_file, keep_header, header_size, keep_newlines)

	if input_is_dir:
		os.chdir(old_pwd)

if __name__ == "__main__":
	parser = argparse.ArgumentParser(description = "Convert docx tables to CSV files.")
	parser.add_argument("-i", metavar = "input_file", type = str,
                    help = "an input directory or file")
	parser.add_argument("-o", metavar = "output_dir", type = str, 
                    help = "an output directory to save CSV files")
	parser.add_argument("-c", action = "store_false",
                    help = "convert all DOCX tables (not only those containing a caption)")
	parser.add_argument("--named_ranges", action = "store_true",
                    help = "produce CSV files for named ranges (XLSX)")
	parser.add_argument("-k", action = "store_true",
                    help = "keep table header in output files (default: false)")
	parser.add_argument("-n", action = "store_true",
                    help = "keep newlines in the table cells (default: false)")
	parser.add_argument("-s", metavar = "header_size", type = int, default = 0,
                    help = "a size of the table header (default: 0)")

	args = parser.parse_args()
	main(input_file = args.i,
		output_dir = args.o,
		use_captions = args.c,
		use_named_ranges = args.named_ranges,
		keep_header = args.k,
		header_size = args.s,
		keep_newlines = args.n)
